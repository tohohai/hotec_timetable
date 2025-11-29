from openpyxl import load_workbook
from .models import (
    Department, TrainingLevel, Major,
    AcademicYear, Semester,
    RoomType, SpecializationGroup, Room,
    Instructor, StudentClass, Subject,
)


def _load_ws(file):
    wb = load_workbook(file, data_only=True)
    return wb.active  # dùng sheet đầu tiên

def import_departments_from_excel(file):
    """
    Excel format:
    A: code   (Mã Khoa)
    B: name   (Tên Khoa)
    """
    ws = _load_ws(file)
    created = updated = 0
    errors = []

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        code, name = row[:2]
        if not code:
            continue
        try:
            obj, is_created = Department.objects.update_or_create(
                code=str(code).strip(),
                defaults={"name": str(name).strip() if name else ""},
            )
            if is_created:
                created += 1
            else:
                updated += 1
        except Exception as e:
            errors.append(f"Dòng {idx}: {e}")

    return created, updated, errors

def import_training_levels_from_excel(file):
    """
    A: code (VD: CD, TC)
    B: name
    C: form (VD: Tập trung, Liên thông...)
    """
    ws = _load_ws(file)
    created = updated = 0
    errors = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        code, name, form = row[:3]
        if not code:
            continue
        try:
            obj, is_created = TrainingLevel.objects.update_or_create(
                code=str(code).strip(),
                defaults={
                    "name": str(name).strip() if name else "",
                    "form": str(form).strip() if form else "Tập trung",
                },
            )
            if is_created: created += 1
            else: updated += 1
        except Exception as e:
            errors.append(f"Dòng {idx}: {e}")
    return created, updated, errors

def import_academic_years_from_excel(file):
    """
    A: code (VD: 2025, 2025-2026...)
    """
    ws = _load_ws(file)
    created = updated = 0
    errors = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        (code,) = row[:1]
        if not code:
            continue
        try:
            obj, is_created = AcademicYear.objects.update_or_create(
                code=str(code).strip(),
                defaults={},
            )
            if is_created: created += 1
            else: updated += 1
        except Exception as e:
            errors.append(f"Dòng {idx}: {e}")
    return created, updated, errors

def import_semesters_from_excel(file):
    """
    A: academic_year_code  (VD: 2025)
    B: code                (VD: HK1)
    C: name                (VD: Học kỳ 1 năm 2025-2026)
    D: start_date          (yyyy-mm-dd)
    E: weeks               (số tuần, VD: 20)
    """
    import datetime
    ws = _load_ws(file)
    created = updated = 0
    errors = []

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        year_code, code, name, start_date, weeks = row[:5]
        if not year_code or not code:
            continue
        try:
            ay = AcademicYear.objects.get(code=str(year_code).strip())
        except AcademicYear.DoesNotExist:
            errors.append(f"Dòng {idx}: Không tìm thấy Năm học '{year_code}'")
            continue

        # start_date có thể là datetime hoặc string
        if isinstance(start_date, str):
            start_date = datetime.datetime.strptime(start_date, "%Y-%m-%d").date()

        try:
            obj, is_created = Semester.objects.update_or_create(
                academic_year=ay,
                code=str(code).strip(),
                defaults={
                    "name": str(name).strip() if name else "",
                    "start_date": start_date,
                    "weeks": int(weeks or 20),
                },
            )
            if is_created: created += 1
            else: updated += 1
        except Exception as e:
            errors.append(f"Dòng {idx}: {e}")

    return created, updated, errors

def import_roomtypes_from_excel(file):
    """
    A: code
    B: name
    """
    ws = _load_ws(file)
    created = updated = 0
    errors = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        code, name = row[:2]
        if not code:
            continue
        try:
            obj, is_created = RoomType.objects.update_or_create(
                code=str(code).strip(),
                defaults={"name": str(name).strip() if name else ""},
            )
            if is_created: created += 1
            else: updated += 1
        except Exception as e:
            errors.append(f"Dòng {idx}: {e}")
    return created, updated, errors

def import_specialization_groups_from_excel(file):
    """
    A: code
    B: name
    """
    ws = _load_ws(file)
    created = updated = 0
    errors = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        code, name = row[:2]
        if not code:
            continue
        try:
            obj, is_created = SpecializationGroup.objects.update_or_create(
                code=str(code).strip(),
                defaults={"name": str(name).strip() if name else ""},
            )
            if is_created: created += 1
            else: updated += 1
        except Exception as e:
            errors.append(f"Dòng {idx}: {e}")
    return created, updated, errors

def import_rooms_from_excel(file):
    """
    A: code
    B: name
    C: room_type_code
    D: specialization_group_codes (mã nhóm chuyên môn, phân cách bằng dấu phẩy)
    """
    ws = _load_ws(file)
    created = updated = 0
    errors = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        code, name, room_type_code, group_codes = row[:4]
        if not code:
            continue
        try:
            rt = RoomType.objects.get(code=str(room_type_code).strip())
        except RoomType.DoesNotExist:
            errors.append(f"Dòng {idx}: Không tìm thấy Loại phòng '{room_type_code}'")
            continue

        try:
            room, is_created = Room.objects.update_or_create(
                code=str(code).strip(),
                defaults={
                    "name": str(name).strip() if name else "",
                    "room_type": rt,
                },
            )
            # cập nhật nhóm chuyên môn (nếu có)
            room.capabilities.clear()
            if group_codes:
                for gcode in str(group_codes).split(","):
                    gcode = gcode.strip()
                    if not gcode:
                        continue
                    try:
                        g = SpecializationGroup.objects.get(code=gcode)
                        room.capabilities.add(g)
                    except SpecializationGroup.DoesNotExist:
                        errors.append(f"Dòng {idx}: Không tìm thấy Nhóm CM '{gcode}'")
            room.save()

            if is_created: created += 1
            else: updated += 1
        except Exception as e:
            errors.append(f"Dòng {idx}: {e}")
    return created, updated, errors

def import_majors_from_excel(file):
    """
    A: code
    B: name
    C: level_code (Bậc: CD/TC...)
    D: department_code (Khoa)
    E: is_general (1/0)
    F: total_credits
    G: total_semesters
    """
    ws = _load_ws(file)
    created = updated = 0
    errors = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        code, name, level_code, dept_code, is_general, total_credits, total_semesters = row[:7]
        if not code:
            continue
        try:
            level = TrainingLevel.objects.get(code=str(level_code).strip())
            dept = Department.objects.get(code=str(dept_code).strip())
        except (TrainingLevel.DoesNotExist, Department.DoesNotExist) as e:
            errors.append(f"Dòng {idx}: {e}")
            continue

        try:
            obj, is_created = Major.objects.update_or_create(
                code=str(code).strip(),
                defaults={
                    "name": str(name).strip() if name else "",
                    "level": level,
                    "department": dept,
                    "is_general": bool(is_general),
                    "total_credits": float(total_credits or 0),
                    "total_semesters": int(total_semesters or 0),
                },
            )
            if is_created: created += 1
            else: updated += 1
        except Exception as e:
            errors.append(f"Dòng {idx}: {e}")
    return created, updated, errors

def import_instructors_from_excel(file):
    """
    A: code
    B: name
    C: department_code
    D: is_leader (1/0)
    E: teaching_quota
    F: admin_quota
    G: conversion_ratio
    """
    ws = _load_ws(file)
    created = updated = 0
    errors = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        code, name, dept_code, is_leader, tq, aq, ratio = row[:7]
        if not code:
            continue
        try:
            dept = Department.objects.get(code=str(dept_code).strip())
        except Department.DoesNotExist:
            errors.append(f"Dòng {idx}: Không tìm thấy Khoa '{dept_code}'")
            continue

        try:
            obj, is_created = Instructor.objects.update_or_create(
                code=str(code).strip(),
                defaults={
                    "name": str(name).strip() if name else "",
                    "department": dept,
                    "is_leader": bool(is_leader),
                    "teaching_quota": float(tq or 415),
                    "admin_quota": float(aq or 480),
                    "conversion_ratio": float(ratio or 3.2),
                },
            )
            if is_created: created += 1
            else: updated += 1
        except Exception as e:
            errors.append(f"Dòng {idx}: {e}")
    return created, updated, errors

def import_student_classes_from_excel(file):
    """
    A: code
    B: name
    C: size
    D: major_code
    E: academic_year_code (Năm nhập học)
    F: homeroom_teacher_code (Mã GV chủ nhiệm - có thể để trống)
    """
    ws = _load_ws(file)
    created = updated = 0
    errors = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        code, name, size, major_code, year_code, gv_code = row[:6]
        if not code:
            continue
        try:
            major = Major.objects.get(code=str(major_code).strip())
            ay = AcademicYear.objects.get(code=str(year_code).strip())
        except (Major.DoesNotExist, AcademicYear.DoesNotExist) as e:
            errors.append(f"Dòng {idx}: {e}")
            continue

        homeroom = None
        if gv_code:
            try:
                homeroom = Instructor.objects.get(code=str(gv_code).strip())
            except Instructor.DoesNotExist:
                errors.append(f"Dòng {idx}: Không tìm thấy GV '{gv_code}' (GVCN), bỏ qua GVCN")

        try:
            obj, is_created = StudentClass.objects.update_or_create(
                code=str(code).strip(),
                defaults={
                    "name": str(name).strip() if name else "",
                    "size": int(size or 0),
                    "major": major,
                    "academic_year": ay,
                    "homeroom_teacher": homeroom,
                },
            )
            if is_created: created += 1
            else: updated += 1
        except Exception as e:
            errors.append(f"Dòng {idx}: {e}")
    return created, updated, errors

def import_subjects_from_excel(file):
    """
    A: code
    B: name
    C: major_code
    D: department_code
    E: required_room_type_code
    F: specialization_group_code
    G: total_periods
    H: form (LT/TH/TH_LT)
    I: semester_index (1,2,3...)
    """
    ws = _load_ws(file)
    created = updated = 0
    errors = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        (
            code, name, major_code, dept_code,
            rt_code, group_code, total_periods,
            form, sem_idx
        ) = row[:9]

        if not code:
            continue

        major = None
        if major_code:
            try:
                major = Major.objects.get(code=str(major_code).strip())
            except Major.DoesNotExist:
                errors.append(f"Dòng {idx}: Không tìm thấy Ngành '{major_code}'")
                major = None

        dept = None
        if dept_code:
            try:
                dept = Department.objects.get(code=str(dept_code).strip())
            except Department.DoesNotExist:
                errors.append(f"Dòng {idx}: Không tìm thấy Khoa '{dept_code}'")

        rt = None
        if rt_code:
            try:
                rt = RoomType.objects.get(code=str(rt_code).strip())
            except RoomType.DoesNotExist:
                errors.append(f"Dòng {idx}: Không tìm thấy Loại phòng '{rt_code}'")

        group = None
        if group_code:
            try:
                group = SpecializationGroup.objects.get(code=str(group_code).strip())
            except SpecializationGroup.DoesNotExist:
                errors.append(f"Dòng {idx}: Không tìm thấy Nhóm CM '{group_code}'")

        # chuẩn hoá form
        form = (form or "").strip()
        if form not in ("LT", "TH", "TH_LT"):
            # bạn có thể mapping từ "MH"/"MĐ" sang LT/TH ở đây nếu muốn
            pass

        try:
            obj, is_created = Subject.objects.update_or_create(
                code=str(code).strip(),
                defaults={
                    "name": str(name).strip() if name else "",
                    "major": major,
                    "department": dept,
                    "required_room_type": rt,
                    "specialization_group": group,
                    "total_periods": float(total_periods or 0),
                    "form": form or "LT",
                    "semester_index": int(sem_idx or 1),
                },
            )
            if is_created: created += 1
            else: updated += 1
        except Exception as e:
            errors.append(f"Dòng {idx}: {e}")

    return created, updated, errors
