from openpyxl import load_workbook
from .models import (
    Department, TrainingLevel, Major,
    AcademicYear, Semester,
    RoomType, SpecializationGroup, Room,
    Instructor, StudentClass, Subject,RoomCapability
)

def _load_ws(file):
    wb = load_workbook(file, data_only=True)
    return wb.active  # dùng sheet đầu tiên


# =============== HÀM TIỆN ÍCH NHẬN FILE HOẶC WORKSHEET ===============

def _get_ws(file_or_ws):
    """
    Nếu truyền vào là file -> mở workbook & lấy sheet đầu.
    Nếu truyền vào là worksheet -> dùng luôn.
    """
    if hasattr(file_or_ws, "iter_rows"):
        return file_or_ws
    return _load_ws(file_or_ws)


# =============== IMPORT TỪNG MODEL (ĐÃ SỬA ĐỂ NHẬN WS) ===============

def import_departments_from_excel(file_or_ws):
    """
    Sheet Departments:
    A: code   (Mã Khoa)
    B: name   (Tên Khoa)
    """
    ws = _get_ws(file_or_ws)
    created = updated = 0
    errors = []

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        code, name = row[:2]
        if not code:
            continue
        try:
            obj, is_created = Department.objects.update_or_create(
                code=str(code).strip(),
                defaults={"name": str(name).strip() if name else ""},
            )
            if is_created:
                created += 1
            else:
                updated += 1
        except Exception as e:
            errors.append(f"Dòng {idx}: {e}")

    return created, updated, errors


def import_training_levels_from_excel(file_or_ws):
    """
    Sheet TrainingLevels:
    A: code (VD: CD, TC)
    B: name
    C: form (VD: Tập trung, Liên thông...)
    """
    ws = _get_ws(file_or_ws)
    created = updated = 0
    errors = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        code, name, form = row[:3]
        if not code:
            continue
        try:
            obj, is_created = TrainingLevel.objects.update_or_create(
                code=str(code).strip(),
                defaults={
                    "name": str(name).strip() if name else "",
                    "form": str(form).strip() if form else "Tập trung",
                },
            )
            if is_created:
                created += 1
            else:
                updated += 1
        except Exception as e:
            errors.append(f"Dòng {idx}: {e}")
    return created, updated, errors


def import_academic_years_from_excel(file_or_ws):
    """
    Sheet AcademicYears:
    A: code (VD: 2025, 2025-2026...)
    """
    ws = _get_ws(file_or_ws)
    created = updated = 0
    errors = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        (code,) = row[:1]
        if not code:
            continue
        try:
            obj, is_created = AcademicYear.objects.update_or_create(
                code=str(code).strip(),
                defaults={},
            )
            if is_created:
                created += 1
            else:
                updated += 1
        except Exception as e:
            errors.append(f"Dòng {idx}: {e}")
    return created, updated, errors


def import_semesters_from_excel(file_or_ws):
    """
    Sheet Semesters:
    A: academic_year_code  (VD: 2025-2026)
    B: code                (VD: HK1)
    C: name                (VD: Học kỳ 1 năm 2025-2026)
    D: start_date          (yyyy-mm-dd)
    E: weeks               (số tuần, VD: 20)
    """
    import datetime
    ws = _get_ws(file_or_ws)
    created = updated = 0
    errors = []

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        year_code, code, name, start_date, weeks = row[:5]
        if not year_code or not code:
            continue
        try:
            ay = AcademicYear.objects.get(code=str(year_code).strip())
        except AcademicYear.DoesNotExist:
            errors.append(f"Dòng {idx}: Không tìm thấy Năm học '{year_code}'")
            continue

        # start_date có thể là datetime hoặc string
        if isinstance(start_date, str):
            try:
                start_date = datetime.datetime.strptime(start_date, "%Y-%m-%d").date()
            except ValueError:
                errors.append(f"Dòng {idx}: Ngày bắt đầu không đúng định dạng yyyy-mm-dd")
                start_date = None

        try:
            obj, is_created = Semester.objects.update_or_create(
                academic_year=ay,
                code=str(code).strip(),
                defaults={
                    "name": str(name).strip() if name else "",
                    "start_date": start_date,
                    "weeks": int(weeks or 20),
                },
            )
            if is_created:
                created += 1
            else:
                updated += 1
        except Exception as e:
            errors.append(f"Dòng {idx}: {e}")

    return created, updated, errors


def import_roomtypes_from_excel(file_or_ws):
    """
    Sheet RoomTypes:
    A: code
    B: name
    """
    ws = _get_ws(file_or_ws)
    created = updated = 0
    errors = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        code, name = row[:2]
        if not code:
            continue
        try:
            obj, is_created = RoomType.objects.update_or_create(
                code=str(code).strip(),
                defaults={"name": str(name).strip() if name else ""},
            )
            if is_created:
                created += 1
            else:
                updated += 1
        except Exception as e:
            errors.append(f"Dòng {idx}: {e}")
    return created, updated, errors


def import_specialization_groups_from_excel(file_or_ws):
    """
    Sheet SpecializationGroups:
    A: code
    B: name
    """
    ws = _get_ws(file_or_ws)
    created = updated = 0
    errors = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        code, name = row[:2]
        if not code:
            continue
        try:
            obj, is_created = SpecializationGroup.objects.update_or_create(
                code=str(code).strip(),
                defaults={"name": str(name).strip() if name else ""},
            )
            if is_created:
                created += 1
            else:
                updated += 1
        except Exception as e:
            errors.append(f"Dòng {idx}: {e}")
    return created, updated, errors


# def import_rooms_from_excel(file_or_ws):
#     """
#     Sheet Rooms:
#     A: code
#     B: name
#     C: room_type_code
#     D: specialization_group_codes (mã nhóm chuyên môn, phân cách bằng dấu phẩy)
#     """
#     ws = _get_ws(file_or_ws)
#     created = updated = 0
#     errors = []
#     for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
#         code, name, room_type_code, group_codes = row[:4]
#         if not code:
#             continue
#         try:
#             rt = RoomType.objects.get(code=str(room_type_code).strip())
#         except RoomType.DoesNotExist:
#             errors.append(f"Dòng {idx}: Không tìm thấy Loại phòng '{room_type_code}'")
#             continue

#         try:
#             room, is_created = Room.objects.update_or_create(
#                 code=str(code).strip(),
#                 defaults={
#                     "name": str(name).strip() if name else "",
#                     "room_type": rt,
#                 },
#             )
#             # cập nhật nhóm chuyên môn (nếu có)
#             room.capabilities.clear()
#             if group_codes:
#                 for gcode in str(group_codes).split(","):
#                     gcode = gcode.strip()
#                     if not gcode:
#                         continue
#                     try:
#                         g = SpecializationGroup.objects.get(code=gcode)
#                         room.capabilities.add(g)
#                     except SpecializationGroup.DoesNotExist:
#                         errors.append(f"Dòng {idx}: Không tìm thấy Nhóm CM '{gcode}'")
#             room.save()

#             if is_created:
#                 created += 1
#             else:
#                 updated += 1
#         except Exception as e:
#             errors.append(f"Dòng {idx}: {e}")
#     return created, updated, errors

def import_rooms_from_excel(file_or_ws):
    """
    Import Phòng học từ Excel.
    Cấu trúc file (6 cột):
    A: code (Mã phòng)
    B: name (Tên phòng)
    C: room_type_code (Mã loại phòng)
    D: capacity (Sức chứa - Số nguyên)
    E: capabilities (Mã nhóm CM, cách nhau dấu phẩy)
    F: allowed_majors (Mã ngành ưu tiên, cách nhau dấu phẩy)
    """
    # Hàm _get_ws giả định bạn đã có sẵn như trong code cũ
    ws = _get_ws(file_or_ws) 
    created = updated = 0
    errors = []

    # Lấy dữ liệu từ dòng 2, giới hạn đọc đúng 6 cột đầu tiên (A->F)
    for idx, row in enumerate(ws.iter_rows(min_row=2, max_col=6, values_only=True), start=2):
        # 1. Unpack đủ 6 giá trị
        # Lưu ý: Nếu file Excel dòng đó thiếu cột, openpyxl sẽ trả về None, cần xử lý kỹ
        if not row or row[0] is None:
            continue
            
        # Đảm bảo row có ít nhất 6 phần tử (đề phòng dòng ngắn)
        row = list(row) + [None]*(6-len(row))
        code, name, room_type_code, capacity_val, group_codes, major_codes = row

        if not code:
            continue

        # --- Xử lý Loại phòng ---
        try:
            rt = RoomType.objects.get(code=str(room_type_code).strip())
        except RoomType.DoesNotExist:
            errors.append(f"Dòng {idx}: Không tìm thấy Loại phòng '{room_type_code}'")
            continue

        # --- Xử lý Sức chứa (Fix lỗi bị 0) ---
        try:
            capacity = int(capacity_val) if capacity_val else 0
        except ValueError:
            capacity = 0 # Nếu nhập chữ hoặc lỗi định dạng thì về 0

        try:
            # --- Tạo hoặc Update Room ---
            room, is_created = Room.objects.update_or_create(
                code=str(code).strip(),
                defaults={
                    "name": str(name).strip() if name else "",
                    "room_type": rt,
                    "capacity": capacity,
                },
            )

            # --- 2. Xử lý Capabilities (RoomCapability - Model trung gian) ---
            # Xóa liên kết cũ của phòng này để cập nhật mới
            RoomCapability.objects.filter(room=room).delete()

            if group_codes:
                # Tách chuỗi bằng dấu phẩy
                for gcode in str(group_codes).split(","):
                    gcode = gcode.strip()
                    if not gcode: continue
                    try:
                        g = SpecializationGroup.objects.get(code=gcode)
                        # Tạo record trong bảng trung gian
                        # Lưu ý: trường trong model bạn là 'group' (không phải specialization_group)
                        RoomCapability.objects.create(room=room, group=g) 
                    except SpecializationGroup.DoesNotExist:
                        errors.append(f"Dòng {idx}: Không tìm thấy Nhóm CM '{gcode}'")

            # --- 3. Xử lý Allowed Majors (Many-to-Many thường) ---
            major_objs = []
            if major_codes:
                for mcode in str(major_codes).split(","):
                    mcode = mcode.strip()
                    if not mcode: continue
                    try:
                        m = Major.objects.get(code=mcode)
                        major_objs.append(m)
                    except Major.DoesNotExist:
                        errors.append(f"Dòng {idx}: Không tìm thấy Ngành '{mcode}'")
            
            # Gán danh sách ngành vào phòng (tự động xóa cũ thay mới)
            room.allowed_majors.set(major_objs)

            # Đếm kết quả
            if is_created:
                created += 1
            else:
                updated += 1

        except Exception as e:
            errors.append(f"Dòng {idx}: Lỗi hệ thống - {str(e)}")

    return created, updated, errors

def import_majors_from_excel(file_or_ws):
    """
    Sheet Majors:
    A: code
    B: name
    C: level_code (Bậc: CD/TC...)
    D: department_code (Khoa)
    E: is_general (1/0)
    F: total_credits
    G: total_semesters
    """
    ws = _get_ws(file_or_ws)
    created = updated = 0
    errors = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        code, name, level_code, dept_code, is_general, total_credits, total_semesters = row[:7]
        if not code:
            continue
        try:
            level = TrainingLevel.objects.get(code=str(level_code).strip())
            dept = Department.objects.get(code=str(dept_code).strip())
        except (TrainingLevel.DoesNotExist, Department.DoesNotExist) as e:
            errors.append(f"Dòng {idx}: {e}")
            continue

        try:
            obj, is_created = Major.objects.update_or_create(
                code=str(code).strip(),
                defaults={
                    "name": str(name).strip() if name else "",
                    "level": level,
                    "department": dept,
                    "is_general": bool(is_general),
                    "total_credits": float(total_credits or 0),
                    "total_semesters": int(total_semesters or 0),
                },
            )
            if is_created:
                created += 1
            else:
                updated += 1
        except Exception as e:
            errors.append(f"Dòng {idx}: {e}")
    return created, updated, errors


def import_instructors_from_excel(file_or_ws):
    """
    Sheet Instructors:
    A: code
    B: name
    C: department_code
    D: is_leader (1/0)
    E: teaching_quota
    F: admin_quota
    G: conversion_ratio
    """
    ws = _get_ws(file_or_ws)
    created = updated = 0
    errors = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        code, name, dept_code, is_leader, tq, aq, ratio = row[:7]
        if not code:
            continue
        try:
            dept = Department.objects.get(code=str(dept_code).strip())
        except Department.DoesNotExist:
            errors.append(f"Dòng {idx}: Không tìm thấy Khoa '{dept_code}'")
            continue

        try:
            obj, is_created = Instructor.objects.update_or_create(
                code=str(code).strip(),
                defaults={
                    "name": str(name).strip() if name else "",
                    "department": dept,
                    "is_leader": bool(is_leader),
                    "teaching_quota": float(tq or 415),
                    "admin_quota": float(aq or 480),
                    "conversion_ratio": float(ratio or 3.2),
                },
            )
            if is_created:
                created += 1
            else:
                updated += 1
        except Exception as e:
            errors.append(f"Dòng {idx}: {e}")
    return created, updated, errors


def import_student_classes_from_excel(file_or_ws):
    """
    Sheet StudentClasses:
    A: code
    B: name
    C: size
    D: major_code
    E: academic_year_code (Năm nhập học)
    F: homeroom_teacher_code (Mã GV chủ nhiệm - có thể để trống)
    """
    ws = _get_ws(file_or_ws)
    created = updated = 0
    errors = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        code, name, size, major_code, year_code, gv_code = row[:6]
        if not code:
            continue
        try:
            major = Major.objects.get(code=str(major_code).strip())
            ay = AcademicYear.objects.get(code=str(year_code).strip())
        except (Major.DoesNotExist, AcademicYear.DoesNotExist) as e:
            errors.append(f"Dòng {idx}: {e}")
            continue

        homeroom = None
        if gv_code:
            try:
                homeroom = Instructor.objects.get(code=str(gv_code).strip())
            except Instructor.DoesNotExist:
                errors.append(f"Dòng {idx}: Không tìm thấy GV '{gv_code}' (GVCN), bỏ qua GVCN")

        try:
            obj, is_created = StudentClass.objects.update_or_create(
                code=str(code).strip(),
                defaults={
                    "name": str(name).strip() if name else "",
                    "size": int(size or 0),
                    "major": major,
                    "academic_year": ay,
                    "homeroom_teacher": homeroom,
                },
            )
            if is_created:
                created += 1
            else:
                updated += 1
        except Exception as e:
            errors.append(f"Dòng {idx}: {e}")
    return created, updated, errors


def import_subjects_from_excel(file_or_ws):
    """
    Sheet Subjects (đÃ CHỈNH THEO MODEL Subject MỚI):

    A: code
    B: name
    C: major_code
    D: managing_department_code
    E: subject_type              (TN/TL/TH/BC/KHAC)
    F: total_periods
    G: max_class_size
    H: required_room_type_code
    I: specialization_group_code
    J: is_external_managed       (1/0)
    K: exam_form                 (TN/TL/TH/BC/KHAC, nếu trống dùng subject_type)
    L: has_separate_marking      (1/0)
    """
    ws = _get_ws(file_or_ws)
    created = updated = 0
    errors = []

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        (
            code, name, major_code, dept_code,
            subject_type, total_periods, max_class_size,
            rt_code, group_code,
            is_external_managed, exam_form, has_separate_marking,
        ) = row[:12]

        if not code:
            continue

        major = None
        if major_code:
            try:
                major = Major.objects.get(code=str(major_code).strip())
            except Major.DoesNotExist:
                errors.append(f"Dòng {idx}: Không tìm thấy Ngành '{major_code}'")

        managing_dept = None
        if dept_code:
            try:
                managing_dept = Department.objects.get(code=str(dept_code).strip())
            except Department.DoesNotExist:
                errors.append(f"Dòng {idx}: Không tìm thấy Khoa '{dept_code}'")

        rt = None
        if rt_code:
            try:
                rt = RoomType.objects.get(code=str(rt_code).strip())
            except RoomType.DoesNotExist:
                errors.append(f"Dòng {idx}: Không tìm thấy Loại phòng '{rt_code}'")

        group = None
        if group_code:
            try:
                group = SpecializationGroup.objects.get(code=str(group_code).strip())
            except SpecializationGroup.DoesNotExist:
                errors.append(f"Dòng {idx}: Không tìm thấy Nhóm CM '{group_code}'")

        subject_type = (subject_type or "").strip() or "KHAC"
        exam_form = (exam_form or "").strip() or subject_type

        try:
            obj, is_created = Subject.objects.update_or_create(
                code=str(code).strip(),
                defaults={
                    "name": str(name).strip() if name else "",
                    "major": major,
                    "managing_department": managing_dept,
                    "subject_type": subject_type,
                    "total_periods": float(total_periods or 0),
                    "max_class_size": int(max_class_size or 35),
                    "required_room_type": rt,
                    "specialization_group": group,
                    "is_external_managed": bool(is_external_managed),
                    "exam_form": exam_form,
                    "has_separate_marking": bool(has_separate_marking),
                },
            )
            if is_created:
                created += 1
            else:
                updated += 1
        except Exception as e:
            errors.append(f"Dòng {idx}: {e}")

    return created, updated, errors



def import_room_capabilities_from_excel(file_or_ws):
    """
    Import Khả năng chuyên môn của Phòng (RoomCapability).
    Sheet Structure (3 cột):
    A: room_code (Mã phòng)
    B: group_code (Mã nhóm CM)
    C: priority (Mức ưu tiên: 1, 2, 3 - Mặc định là 1)
    """
    ws = _get_ws(file_or_ws) # Hàm helper lấy worksheet
    created = updated = 0
    errors = []

    # Duyệt từ dòng 2, đọc 3 cột đầu tiên
    for idx, row in enumerate(ws.iter_rows(min_row=2, max_col=3, values_only=True), start=2):
        # Đảm bảo row có đủ độ dài để tránh lỗi index
        row = list(row) + [None] * (3 - len(row))
        
        r_code, g_code, priority_val = row

        # Nếu thiếu mã phòng hoặc mã nhóm thì bỏ qua
        if not r_code or not g_code:
            continue

        # 1. Tìm Phòng (Room)
        try:
            room_obj = Room.objects.get(code=str(r_code).strip())
        except Room.DoesNotExist:
            errors.append(f"Dòng {idx}: Không tìm thấy Phòng mã '{r_code}'")
            continue

        # 2. Tìm Nhóm chuyên môn (SpecializationGroup)
        try:
            group_obj = SpecializationGroup.objects.get(code=str(g_code).strip())
        except SpecializationGroup.DoesNotExist:
            errors.append(f"Dòng {idx}: Không tìm thấy Nhóm CM mã '{g_code}'")
            continue

        # 3. Xử lý Mức ưu tiên (Priority)
        # Mặc định là 1 nếu dữ liệu lỗi hoặc trống
        try:
            p = int(priority_val) if priority_val else 1
            if p not in [1, 2, 3]: # Nếu nhập số khác 1,2,3 thì ép về 1
                p = 1
        except ValueError:
            p = 1

        try:
            # 4. Lưu vào Database
            # Dùng update_or_create dựa trên cặp khóa (room, group)
            # Nếu đã tồn tại cặp này -> cập nhật priority
            # Nếu chưa -> tạo mới
            cap, is_created = RoomCapability.objects.update_or_create(
                room=room_obj,
                group=group_obj,
                defaults={
                    "priority": p
                }
            )

            if is_created:
                created += 1
            else:
                updated += 1

        except Exception as e:
            errors.append(f"Dòng {idx}: Lỗi hệ thống - {str(e)}")

    return created, updated, errors
# =============== IMPORT ALL TỪ 1 FILE NHIỀU SHEET ===============

def import_all_from_excel(file):
    """
    Import tất cả danh mục từ 1 file Excel nhiều sheet.
    Thứ tự sheet (phụ thuộc FK):
      1. Departments
      2. TrainingLevels
      3. AcademicYears
      4. Semesters
      5. RoomTypes
      6. SpecializationGroups
      7. Rooms
      8. Majors
      9. Instructors
      10. StudentClasses
      11. Subjects
    """
    wb = load_workbook(file, data_only=True)

    sheet_importers = [
        ("Departments",          import_departments_from_excel),
        ("TrainingLevels",       import_training_levels_from_excel),
        ("AcademicYears",        import_academic_years_from_excel),
        ("Semesters",            import_semesters_from_excel),
        ("RoomTypes",            import_roomtypes_from_excel),
        ("SpecializationGroups", import_specialization_groups_from_excel),
        ("Rooms",                import_rooms_from_excel),
        ("Majors",               import_majors_from_excel),
        ("Instructors",          import_instructors_from_excel),
        ("StudentClasses",       import_student_classes_from_excel),
        ("Subjects",             import_subjects_from_excel),
    ]

    overall_result = {}

    for sheet_name, importer in sheet_importers:
        if sheet_name not in wb.sheetnames:
            overall_result[sheet_name] = {
                "created": 0,
                "updated": 0,
                "errors": [f"Sheet '{sheet_name}' không tồn tại, bỏ qua."],
            }
            continue

        ws = wb[sheet_name]
        created, updated, errors = importer(ws)

        overall_result[sheet_name] = {
            "created": created,
            "updated": updated,
            "errors": errors,
        }

    return overall_result
